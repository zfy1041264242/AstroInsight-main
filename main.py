import os
from random import random
from app.utils.tool import get_related_keyword,remove_number_prefix,extract_hypothesis,search_releated_paper,extract_technical_entities,extract_message,paper_compression,review_mechanism,extract_message_review,extract_message_review_moa
from app.utils.arxiv_api import search_paper
from app.utils.llm_api import call_with_deepseek,call_with_qwenmax
from app.core.prompt import fact_extraction_prompt,hypothesis_generate_prompt,hypotheses_prompt,initial_idea_prompt,technical_optimizatio_prompt,MoA_based_optimization_prompt,human_ai_collaboration_prompt
from openpyxl import load_workbook,Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from app.core.moa import moa_idea_iteration,moa_table
from app.core.config import OUTPUT_PATH

def process_paper(paper):
    if 'abstract' not in paper:
        return None

    user_prompt = f"""Now,please following these rules to extract the factual information from following paper:\ntitile:{paper['title']}\nabstract:\n{paper['abstract']}\n"""

    result = call_with_deepseek(system_prompt=fact_extraction_prompt(), question=user_prompt)

    return {
        "topic": paper['topic'],
        "title": paper['title'],
        "abstract": paper['abstract'],
        "result": result
    }

def Fact_Information_Extraction(Keyword,SearchPaperNum,  user_id, task):
    Keywords = get_related_keyword(Keyword=Keyword)
    Keywords.append(Keyword)
    papers=search_paper(Keywords=Keywords, Limit=SearchPaperNum)

    print(f"\033[1;32m | INFO     | fact information extraction... \033[0m")
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task.id}/{Keyword}"
    # 加载现有的 XLSX 文件
    file_path = fr"{file_path_prefix}/fact_information_{Keyword}.xlsx"
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    start_row = ws.max_row + 1  # 在最后一行之后开始追加

    results = []
    # 并发执行
    with ThreadPoolExecutor(max_workers=5) as executor:  # max_workers 可调节
        future_to_paper = {executor.submit(process_paper, paper): paper for paper in papers}
        for future in as_completed(future_to_paper):
            res = future.result()
            if res:
                results.append(res)

    # 写入 Excel
    for res in results:
        ws.cell(row=start_row, column=1, value=res["topic"])
        ws.cell(row=start_row, column=2, value=res["title"])
        ws.cell(row=start_row, column=3, value=res["abstract"])
        ws.cell(row=start_row, column=4, value=res["result"])
        start_row += 1

    # 保存
    wb.save(file_path)
    print(f"\033[1;32m | INFO     | fact information extraction:ok! \033[0m")

    return file_path,Keywords

def Hypothesis_Generate(Keyword,Fact_File_Path,Keywords,shuffle=False,random_num=5, user_id="", task=None):
    print(f"\033[1;32m | INFO     | hypothesis_generate... \033[0m")

    wb = load_workbook(filename=Fact_File_Path)
    ws = wb.active

    fact_information = []

    for keyword in Keywords:
        index=0
        for paper in ws.iter_rows(min_row=2,values_only=False):
            if paper[0].value==keyword and index<random_num:
                temp=remove_number_prefix(paper[3].value)
                fact_information += temp.split('\n')
                index+=1

    if shuffle==True:
        random.shuffle(fact_information)

    Known_Information=""
    index=0
    for information in fact_information:
        if information!="" and information!="\n" :
            Known_Information+=f"{index+1}. {information}\n"
            index+=1

    print(f"\033[1;32m | INFO     | the fact information is :\n{Known_Information} \033[0m")

    user_prompt = hypothesis_generate_prompt(Keyword=Keyword,Known_Information=Known_Information)

    result = call_with_deepseek(system_prompt="You are a research expert.", question=user_prompt, temperature=1.5)
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task.id}/{Keyword}/Hypotheses"
    os.makedirs(file_path_prefix, exist_ok=True)
    with open(fr"{file_path_prefix}/fromfact_{Keyword}_input.md", 'w', encoding='utf-8') as f: f.write(user_prompt)
    with open(fr"{file_path_prefix}/fromfact_{Keyword}_result.md", 'w', encoding='utf-8') as f: f.write(result)

    print(f"\033[1;32m | INFO     | hypothesis_generate:ok! \033[0m")

    return fr"{file_path_prefix}/fromfact_{Keyword}_result.md"

def Initial_Idea(Keyword,SearchPaperNum=5,Compression=True, user_id="",task=None):
    fact_file_path,Keywords = Fact_Information_Extraction(Keyword=Keyword,SearchPaperNum=SearchPaperNum, user_id=user_id, task=task)

    hypo_file_path = Hypothesis_Generate(Keyword=Keyword,Fact_File_Path=fact_file_path, Keywords=Keywords, user_id=user_id, task=task)

    hypotheses = extract_hypothesis(file=hypo_file_path, split_section="Hypothesis")  # getting Hypothesis

    hypotheses_index,hypotheses_result = hypotheses_prompt(Hypotheses=hypotheses)

    keynum, relatedPaper, keyword_str = search_releated_paper(topic=Keyword, max_paper_num=SearchPaperNum, compression=Compression, user_id=user_id, task=task)

    title_abstract_prompt = ""

    if Compression==True:
        for index,paper in enumerate(relatedPaper):
            title_abstract_prompt+=f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"""
    else:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"""

    user_prompt = initial_idea_prompt(hypotheses_prompt=hypotheses_result, title_abstract_prompt=title_abstract_prompt,keyword_str=keyword_str,hypotheses_index=hypotheses_index,index=index,keynum=keynum)
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task.id}/{Keyword}/Idea"

    os.makedirs(file_path_prefix, exist_ok=True)
    with open(fr"{file_path_prefix}/{Keyword}_initial_input.md",'w', encoding='utf-8') as f:
        f.write(user_prompt)

    # initial_idea_result = call_with_deepseek(question=user_prompt,temperature=1.5)
    initial_idea_result = call_with_qwenmax(question=user_prompt)

    with open(fr"{file_path_prefix}/{Keyword}_initial_result.md",'w', encoding='utf-8') as f:
        f.write(initial_idea_result)

    print(f"\033[1;32m | INFO     | the initial idea draft ok \033[0m")

    return fr"{file_path_prefix}/{Keyword}_initial_result.md"

def Technical_Optimization(Keyword,Initial_Idea_Result_File,Compression=True, user_id="",task=None):
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task.id}/{Keyword}/Idea"
    technical_keywords, Initial_Idea_Result = extract_technical_entities(Initial_Idea_Result_File,split_section="Paper Abstract")

    text, target_paper_title = extract_message(Initial_Idea_Result_File, split_section="Paper Title")

    print(f"\033[1;32m | INFO     | retrieve papers related to technical entities... \033[0m")

    readysearch_key=[]

    for index,keyword in enumerate(technical_keywords):
        if index > 3:
            break
        readysearch_key.append(keyword['entity'])

    Papers = search_paper(Keywords=readysearch_key, Limit=2)  # getting technical papers

    title_abstract_prompt=""
    relatedPaper = []

    for paper in Papers:
        if Compression==True:
            try:
                print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} \033[0m")
                compression_result = paper_compression(doi=paper["doi"],title=paper["title"],topic=Keyword, user_id=user_id, task=task)
                print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} State: OK! \033[0m")
            except Exception as e:
                print(f"\033[1;31m | ERRO     | compressed paper information: {paper['title']} State: Miss! \033[0m")
                print(f"\033[1;31m | ERROR    | Error details: {str(e)} \033[0m")
                print(f"\033[1;31m | ERROR    | Error type: {type(e).__name__} \033[0m")
                compression_result="None"
        else: compression_result="None"

        try:
            relatedPaper.append({
                "title": paper["title"],
                "abstract": paper["abstract"],
                "compression_result": compression_result
            })
        except:
            pass

    if Compression:
        for index,paper in enumerate(relatedPaper):
            title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"""
    else:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"""

    user_prompt = technical_optimizatio_prompt(title=target_paper_title, title_abstract_prompt=title_abstract_prompt ,Initial_Idea_Result=Initial_Idea_Result)

    with open(fr"{file_path_prefix}/{Keyword}_technical_optimization_input.md", 'w', encoding='utf-8') as f:
        f.write(user_prompt)

    idea_iteration_result = call_with_qwenmax(question=user_prompt)

    with open(fr"{file_path_prefix}/{Keyword}_technical_ optimization_result.md", 'w', encoding='utf-8') as f:
        f.write(idea_iteration_result)

    print(f"\033[1;32m | INFO     | iteration Technical_Optimization ok \033[0m")

    return fr"{file_path_prefix}/{Keyword}_technical_ optimization_result.md"

def MoA_Based_Optimization(Keyword, Technical_Optimization_Result_File,Compression=True, user_id="",task=None):
    print(f"\033[1;32m | INFO     |  start idea draft iter MoA based optimization... \033[0m")
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task.id}/{Keyword}/Idea"
    file_path_review_prefix = fr"{OUTPUT_PATH}/{user_id}/{task.request.id}/{Keyword}/Review"
    draft, target_paper_title = extract_message(file=Technical_Optimization_Result_File, split_section="Paper Title") #getting initial draft title

    information = review_mechanism(topic=Keyword, draft=draft, user_id=user_id, task=task)

    text, next_optimization = extract_message_review(
        file=fr"{file_path_review_prefix}/{Keyword}_review.md",
        split_section="Next Steps for Optimization") #getting review

    target_next_optimization=""
    for index,opt in enumerate(next_optimization.values()):
        target_next_optimization+=f"\n{index+1}.{opt}"

    print(f"\033[1;32m | INFO     |  next optimization target is:\n {target_next_optimization} \033[0m")

    readysearch_key=[]
    optimization_keywords=""

    for keyword in information:
        optimization_keywords+=f"\n{keyword['keyword']}"

    print(readysearch_key)

    Papers = search_paper(Keywords=readysearch_key, Limit=2)  # getting technical papers

    title_abstract_prompt=""
    relatedPaper = []

    for paper in Papers:
        if Compression==True:
            try:
                print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} \033[0m")
                compression_result = paper_compression(doi=paper["doi"],title=paper["title"],topic=Keyword, user_id=user_id, task=task)
                print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} State: OK! \033[0m")
            except Exception as e:
                print(f"\033[1;31m | ERRO     | compressed paper information: {paper['title']} State: Miss! \033[0m")
                print(f"\033[1;31m | ERROR    | Error details: {str(e)} \033[0m")
                print(f"\033[1;31m | ERROR    | Error type: {type(e).__name__} \033[0m")
                compression_result="None"
        else: compression_result="None"

        try:
            relatedPaper.append({
                "title":paper["title"],
                "abstract":paper["abstract"],
                "compression_result": compression_result
            })
        except:
            pass

    if Compression==True:
        for index,paper in enumerate(relatedPaper):
            title_abstract_prompt+=f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"""
    else:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"""

    user_prompt = MoA_based_optimization_prompt(target_next_optimization=target_next_optimization,optimization_keywords=optimization_keywords,title_abstract_prompt=title_abstract_prompt,draft=draft)

    with open(fr"{file_path_prefix}/{Keyword}_moa_input.md", 'w', encoding='utf-8') as f:
        f.write(user_prompt)

    idea_iteration_result = moa_idea_iteration(topic=Keyword, user_prompt=user_prompt, user_id=user_id, task=task)

    with open(fr"{file_path_prefix}/{Keyword}_moa_result.md", 'w', encoding='utf-8') as f:
        f.write(idea_iteration_result)

    print(f"\033[1;32m | INFO     | iteration MoA based optimization ok \033[0m")
    return fr"{file_path_prefix}/{Keyword}_moa_result.md"


def Human_AI_Collaboration(Keyword, MoA_Based_Optimization_Result_File, Compression=True, user_id="", task=None):
    print(f"\033[1;32m | INFO     |  start idea draft iter Human_AI_Collaboration... \033[0m")
    file_path_prefix = fr"{OUTPUT_PATH}/{user_id}/{task.id}/{Keyword}/MOA"
    file_path_result_prefix = fr"{OUTPUT_PATH}/{user_id}/{task.id}/{Keyword}/Idea"

    idea_draft, target_paper_title = extract_message(file=MoA_Based_Optimization_Result_File, split_section="Paper Title")

    print(idea_draft.split("### Summary of the Differences in This Iteration:")[0].strip())
    moa_table(topic=Keyword, draft=idea_draft.split("### Summary of the Differences in This Iteration:")[0].strip(), user_id=user_id, task=task)

    _, next_optimization = extract_message_review_moa(file=fr"{file_path_prefix}/{Keyword}_review_moa.md", split_section="Overall Opinions")

    text, optimize_messages = extract_message_review(file=fr"{file_path_prefix}/{Keyword}_review_moa.md", split_section="Iterative Optimization Search Keywords")

    keywords = []

    for optimize_message in optimize_messages.values():
        keywords.append({'keyword': optimize_message})

    target_next_optimization = ""
    for opt in next_optimization:
        target_next_optimization += f"\n{opt}"

    print(f"\033[1;32m | INFO     |  next optimization target is :\n{target_next_optimization} \033[0m")

    readysearch_key = []
    optimization_keywords = ""

    for keyword in keywords:
        readysearch_key.append(keyword['keyword'])
        optimization_keywords += f"\n{keyword['keyword']}"

    print(f"\033[1;32m | INFO     |  ready search key :\n{readysearch_key} \033[0m")

    Papers = search_paper(Keywords=readysearch_key, Limit=2)  # getting technical papers

    title_abstract_prompt=""
    relatedPaper = []

    for paper in Papers:
        if Compression:
            try:
                print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} \033[0m")
                compression_result = paper_compression(doi=paper["doi"],title=paper["title"],topic=Keyword, user_id=user_id, task=task)
                print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} State: OK! \033[0m")
            except Exception as e:
                print(f"\033[1;31m | ERRO     | compressed paper information: {paper['title']} State: Miss! \033[0m")
                print(f"\033[1;31m | ERROR    | Error details: {str(e)} \033[0m")
                print(f"\033[1;31m | ERROR    | Error type: {type(e).__name__} \033[0m")
                compression_result="None"
        else: compression_result="None"

        try:
            relatedPaper.append({
                "title":paper["title"],
                "abstract":paper["abstract"],
                "compression_result": compression_result
            })
        except:
            pass

    if Compression:
        for index,paper in enumerate(relatedPaper):
            title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"""
    else:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"""

    user_prompt = human_ai_collaboration_prompt(target_next_optimization=target_next_optimization,optimization_keywords=optimization_keywords,title_abstract_prompt=title_abstract_prompt,draft=idea_draft)

    with open(fr"{file_path_result_prefix}/{Keyword}_human_ai_collaboration_input.md",'w', encoding='utf-8') as f:
        f.write(user_prompt)

    idea_iteration_result = call_with_qwenmax(system_prompt="You are a research assistant.", question=user_prompt)


    with open(fr"{file_path_result_prefix}/{Keyword}_human_ai_collaboration_result.md", 'w', encoding='utf-8') as f:
        f.write(idea_iteration_result)

    print(f"\033[1;32m | INFO     | iteration Human_AI_Collaboration ok \033[0m")

    return fr"{file_path_result_prefix}/{Keyword}_human_ai_collaboration_result.md"


def main(task, user_id, Keyword, SearchPaperNum):
    # 在代码开头删除代理环境变量
    # if 'http_proxy' in os.environ:
    #     del os.environ['http_proxy']
    # if 'https_proxy' in os.environ:
    #     del os.environ['https_proxy']
    # if 'HTTP_PROXY' in os.environ:
    #     del os.environ['HTTP_PROXY']
    # if 'HTTPS_PROXY' in os.environ:
    #     del os.environ['HTTPS_PROXY']
    # print(os.environ)
    steps = []
    steps.append("Start Initial_Idea")
    # task.update_state(state='STEP_1', meta={"data:": steps})
    Initial_Idea_Result_File = Initial_Idea(Keyword=Keyword, SearchPaperNum=SearchPaperNum, Compression=True, user_id=user_id, task=task)
    steps.append("End Initial_Idea")
    # task.update_state(state='STEP_2', meta={"data:": steps})

    steps.append("Start Technical_Optimization")
    # task.update_state(state='STEP_3', meta={"data:": steps})
    Technical_Optimization_Result_File = Technical_Optimization(Keyword=Keyword, Initial_Idea_Result_File=Initial_Idea_Result_File, Compression=True, user_id=user_id, task=task)
    steps.append("End Technical_Optimization")
    # task.update_state(state='STEP_4', meta={"data:": steps})

    steps.append("Start MoA_Based_Optimization")
    # task.update_state(state='STEP_5', meta={"data:": steps})
    MoA_Based_Optimization_Result_File = MoA_Based_Optimization(Keyword=Keyword, Technical_Optimization_Result_File=Technical_Optimization_Result_File, Compression=True, user_id=user_id, task=task)
    steps.append("End MoA_Based_Optimization")
    # task.update_state(state='STEP_6', meta={"data:": steps})

    steps.append("Start Human_AI_Collaboration")
    # task.update_state(state='STEP_7', meta={"data:": steps})
    file_path = Human_AI_Collaboration(Keyword=Keyword, MoA_Based_Optimization_Result_File=MoA_Based_Optimization_Result_File, Compression=True, user_id=user_id, task=task)
    steps.append("End Human_AI_Collaboration")
    # task.update_state(state='STEP_8', meta={"data:": steps})
    print("--------------------------over---------------------------")
    return {
        "data": steps,
        "result": file_path
    }



if __name__ == "__main__":
    # 定义一个类
    class Task:
        def __init__(self, id):
            self.id = id  # 实例属性


    # 创建对象并使用属性
    task = Task("123456")
    main(Keyword="database performance optimization", SearchPaperNum=2, user_id="123", task=task)
